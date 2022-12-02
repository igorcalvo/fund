from os import path
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from pandas import DataFrame
from re import match, I

class ValueInputOption():
    RAW = 'RAW'
    USER_ENTERED = 'USER_ENTERED'

def authenticate() -> Credentials:
    scopes = ['https://www.googleapis.com/auth/spreadsheets']
    creds = None
    folder = 'sheets_auth'

    if path.exists(f'{folder}/token.json'):
        creds = Credentials.from_authorized_user_file(f'{folder}/token.json', scopes)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(f'{folder}/credentials.json', scopes)
            creds = flow.run_local_server(port=0)

        with open(f'{folder}/token.json', 'w') as token:
            token.write(creds.to_json())
    return creds

def df_from_sheet(sheet_values: list) -> DataFrame:
    return DataFrame(sheet_values[1:], columns=sheet_values[0])

def read_sheet(sheet_id: str, sheet_name: str, sheet_range: str = '') -> DataFrame:
    creds = authenticate()
    service = build('sheets', 'v4', credentials=creds)
    sheets = service.spreadsheets()

    range = sheet_name if sheet_range == '' else f'{sheet_name}!{sheet_range}'
    result = sheets.values().get(spreadsheetId=sheet_id, range=range).execute()
    values = result.get('values', [])

    if not values:
        print(f'read_sheet: no data found for: id "{sheet_id}" and sheet "{sheet_name}"')
        return None

    df = df_from_sheet(values)
    return df

def df_to_list_of_lists(df: DataFrame) -> list:
    return df.values.tolist()

def create_sheet_tab(sheet_id: str, sheet_name: str):
    creds = authenticate()
    service = build('sheets', 'v4', credentials=creds)
    body = {'requests': [{'addSheet': {'properties': {'title': sheet_name}}}]}
    request = service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body=body)
    response = request.execute()
    print(response)

def split_cell(cell: str) -> tuple:
    regex_match = match(r"([a-z]+)([0-9]+)", cell, I)
    if regex_match:
        items = regex_match.groups()
        return items
    else:
        raise Exception(f'split_cell error parsing cell "{cell}": {e}')

def calculate_final_cell(df: DataFrame, starting_cell: str) -> str:
    starting_col, starting_row_index = split_cell(starting_cell)
    starting_col_index = column_index_from_string(starting_col)

    ending_row_index = int(starting_row_index) + df.shape[0]
    ending_col_index = starting_col_index + df.shape[1] - 1
    ending_col = get_column_letter(ending_col_index)

    result = f'{ending_col}{ending_row_index}'
    return result

def write_sheet(sheet_id: str, sheet_name: str, df: DataFrame, first_cell_offset: str = ''):
    creds = authenticate()
    service = build('sheets', 'v4', credentials=creds)
    values_list = df_to_list_of_lists(df)
    body = {'values': values_list}
    value_input_option = ValueInputOption.USER_ENTERED
    updated_range = sheet_name if offset == '' else f'{sheet_name}!{first_cell_offset}:{calculate_final_cell(df, first_cell_offset)}'

    result = {}
    try:
        result = service.spreadsheets().values().update(spreadsheetId=sheet_id, range=updated_range, valueInputOption=value_input_option, body=body).execute()
    except Exception as e:
        if f'"Unable to parse range: {updated_range}"' in str(e):
            create_sheet_tab(sheet_id, sheet_name)
            result = service.spreadsheets().values().update(spreadsheetId=sheet_id, range=updated_range, valueInputOption=value_input_option, body=body).execute()
        else:
            raise(e)
    finally:
        if 'updatedCells' in result.keys():
            print(f"write_sheet: {result.get('updatedCells')} cells updated.")
    return result
